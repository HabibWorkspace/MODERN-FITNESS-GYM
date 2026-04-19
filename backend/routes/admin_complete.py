"""Complete admin routes for member and trainer management."""
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from models.user import User, UserRole
from models.member_profile import MemberProfile
from models.trainer_profile import TrainerProfile
from models.package import Package
from models.transaction import Transaction, TransactionStatus, TransactionType
from services.password_service import PasswordService
from middleware.rbac import require_admin
from database import db
from datetime import datetime, timedelta
from sqlalchemy import func
import uuid
from io import BytesIO

admin_complete_bp = Blueprint('admin_complete', __name__)

# ============================================================================
# MEMBER ROUTES
# ============================================================================

@admin_complete_bp.route('/members', methods=['GET'])
@require_admin
def list_members():
    """List all members with pagination."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    paginated = MemberProfile.query.paginate(page=page, per_page=per_page, error_out=False)
    
    members = []
    for member in paginated.items:
        member_dict = member.to_dict()
        user = User.query.get(member.user_id)
        if user:
            member_dict['username'] = user.username
        
        # Add trainer information if assigned
        if member.trainer_id:
            trainer = TrainerProfile.query.get(member.trainer_id)
            if trainer:
                member_dict['trainer_name'] = trainer.full_name
                member_dict['trainer_specialization'] = trainer.specialization
            else:
                member_dict['trainer_name'] = None
        else:
            member_dict['trainer_name'] = None
        
        members.append(member_dict)
    
    return jsonify({
        'members': members,
        'total': paginated.total,
        'page': page,
        'per_page': per_page
    }), 200


@admin_complete_bp.route('/members/export', methods=['GET'])
@require_admin
def export_members_excel():
    """Export all members to Excel file."""
    try:
        # Lazy import openpyxl only when needed
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Get all members sorted by member_number
        all_members = MemberProfile.query.order_by(MemberProfile.member_number).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Members List"
        
        # Define headers (only essential columns)
        headers = [
            'Member ID', 'Full Name', 'Phone', 'Email', 'Gender', 
            'Date of Birth', 'Admission Date', 'Current Package', 'Trainer', 'Status'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        header_font = Font(bold=True, size=12, color='000000')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add title row
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = 'MODERN FITNESS GYM - MEMBERS LIST'
        title_cell.font = Font(bold=True, size=16, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        ws.row_dimensions[1].height = 30
        
        # Add headers in row 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # Add data rows starting from row 3
        for row_num, member in enumerate(all_members, 3):
            # Get related data
            package = Package.query.get(member.current_package_id) if member.current_package_id else None
            package_name = package.name if package else 'Not Assigned'
            
            trainer = TrainerProfile.query.get(member.trainer_id) if member.trainer_id else None
            trainer_name = trainer.full_name if trainer else 'Not Assigned'
            
            # Format dates
            dob = member.date_of_birth.strftime('%d-%b-%Y') if member.date_of_birth else 'N/A'
            admission = member.admission_date.strftime('%d-%b-%Y') if member.admission_date else 'N/A'
            
            # Determine status
            status = 'Frozen' if member.is_frozen else 'Active'
            
            # Write row data
            row_data = [
                member.member_number or 'N/A',
                member.full_name or 'N/A',
                member.phone or 'N/A',
                member.email or 'N/A',
                member.gender or 'N/A',
                dob,
                admission,
                package_name,
                trainer_name,
                status
            ]
            
            # Alternate row colors
            row_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid') if row_num % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
                cell.fill = row_fill
                cell.font = Font(size=11)
        
        # Set column widths
        column_widths = {
            'A': 12,  # Member ID
            'B': 25,  # Full Name
            'C': 15,  # Phone
            'D': 25,  # Email
            'E': 10,  # Gender
            'F': 15,  # DOB
            'G': 15,  # Admission Date
            'H': 20,  # Package
            'I': 20,  # Trainer
            'J': 12   # Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        filename = f"Modern_Fitness_Members_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/export', methods=['POST'])
@require_admin
def export_finance_excel():
    """Export finance transactions to Excel file."""
    try:
        # Lazy import openpyxl only when needed
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        data = request.get_json()
        transactions = data.get('transactions', [])
        selected_month = data.get('month', '')
        
        if not transactions:
            return jsonify({'error': 'No transactions to export'}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Report"
        
        # Define headers - simplified (removed Phone, Status, Type)
        headers = [
            'Member ID', 'Member Name', 'Amount (Rs.)', 
            'Due Date', 'Paid Date'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        header_font = Font(bold=True, size=12, color='000000')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add title row
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        month_text = ""
        if selected_month:
            try:
                month_date = datetime.strptime(selected_month + '-01', '%Y-%m-%d')
                month_text = f" - {month_date.strftime('%B %Y')}"
            except:
                pass
        title_cell.value = f'MODERN FITNESS GYM - FINANCE REPORT{month_text}'
        title_cell.font = Font(bold=True, size=16, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        ws.row_dimensions[1].height = 30
        
        # Add headers in row 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # Add data rows starting from row 3
        total_amount = 0
        paid_count = 0
        
        for row_num, txn in enumerate(transactions, 3):
            # Format dates
            due_date = datetime.fromisoformat(txn['due_date'].replace('Z', '')).strftime('%d-%b-%Y') if txn.get('due_date') else 'N/A'
            paid_date = datetime.fromisoformat(txn['paid_date'].replace('Z', '')).strftime('%d-%b-%Y') if txn.get('paid_date') else 'N/A'
            
            # Write row data
            row_data = [
                txn.get('member_id', 'N/A'),
                txn.get('member_name', 'N/A'),
                float(txn.get('amount', 0)),
                due_date,
                paid_date
            ]
            
            # Alternate row colors
            row_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid') if row_num % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            # Highlight paid rows (has paid_date)
            if paid_date != 'N/A':
                row_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                total_amount += float(txn.get('amount', 0))
                paid_count += 1
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='left' if col_num != 3 else 'right', vertical='center')
                cell.border = border
                cell.fill = row_fill
                cell.font = Font(size=11)
                
                # Format amount as currency
                if col_num == 3:
                    cell.number_format = '#,##0.00'
        
        # Add summary rows
        summary_row = len(transactions) + 3
        
        # Total Collected row
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        total_label_cell = ws[f'A{summary_row}']
        total_label_cell.value = 'TOTAL COLLECTED'
        total_label_cell.font = Font(bold=True, size=12, color='000000')
        total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_label_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        total_label_cell.border = border
        
        total_amount_cell = ws[f'C{summary_row}']
        total_amount_cell.value = total_amount
        total_amount_cell.font = Font(bold=True, size=12, color='000000')
        total_amount_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_amount_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        total_amount_cell.border = border
        total_amount_cell.number_format = '#,##0.00'
        
        # Paid Count row
        summary_row += 1
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        count_label_cell = ws[f'A{summary_row}']
        count_label_cell.value = 'PAYMENTS RECEIVED'
        count_label_cell.font = Font(bold=True, size=11, color='000000')
        count_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        count_label_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        count_label_cell.border = border
        
        count_cell = ws[f'C{summary_row}']
        count_cell.value = paid_count
        count_cell.font = Font(bold=True, size=11, color='000000')
        count_cell.alignment = Alignment(horizontal='right', vertical='center')
        count_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        count_cell.border = border
        
        # Set column widths
        column_widths = {
            'A': 12,  # Member ID
            'B': 30,  # Member Name
            'C': 18,  # Amount
            'D': 18,  # Due Date
            'E': 18   # Paid Date
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        month_suffix = f"_{selected_month}" if selected_month else ""
        filename = f"Finance_Report{month_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>', methods=['GET'])
@require_admin
def get_member(member_id):
    """Get a specific member by ID."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    member_dict = member.to_dict()
    user = User.query.get(member.user_id)
    if user:
        member_dict['username'] = user.username
    
    # Add trainer information if assigned
    if member.trainer_id:
        trainer = TrainerProfile.query.get(member.trainer_id)
        if trainer:
            member_dict['trainer_name'] = trainer.full_name
            member_dict['trainer_specialization'] = trainer.specialization
        else:
            member_dict['trainer_name'] = None
    else:
        member_dict['trainer_name'] = None
    
    return jsonify(member_dict), 200


@admin_complete_bp.route('/members', methods=['POST'])
@require_admin
def create_member():
    """Create a new member."""
    data = request.get_json()
    
    # Validate required fields (only full_name and phone are required)
    required_fields = ['full_name', 'phone']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    try:
        # Create user with auto-generated username
        # Use email if provided, otherwise use phone
        if data.get('email'):
            auto_username = data['email'].split('@')[0] + '_' + str(uuid.uuid4())[:8]
        else:
            auto_username = 'member_' + data['phone'][-4:] + '_' + str(uuid.uuid4())[:8]
        auto_password = str(uuid.uuid4())
        
        user = User(
            username=auto_username,
            password_hash=PasswordService.hash_password(auto_password),
            role=UserRole.MEMBER,
            is_active=True
        )
        db.session.add(user)
        db.session.flush()
        
        # Parse date_of_birth if provided
        dob = None
        if data.get('date_of_birth'):
            try:
                from datetime import datetime as dt
                dob = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except:
                dob = None
        
        # Parse admission_date if provided
        admission_date = datetime.utcnow()
        if data.get('admission_date'):
            try:
                from datetime import datetime as dt
                admission_date = dt.strptime(data['admission_date'], '%Y-%m-%d')
            except:
                admission_date = datetime.utcnow()
        
        # Calculate package dates if package is assigned
        package_start_date = None
        package_expiry_date = None
        if data.get('package_id'):
            package = Package.query.get(data['package_id'])
            if package:
                package_start_date = datetime.utcnow()
                package_expiry_date = package_start_date + timedelta(days=package.duration_days)
        
        # Auto-generate member_number starting from 10
        # Get the highest member_number and increment by 1
        max_member_number = db.session.query(func.max(MemberProfile.member_number)).scalar()
        if max_member_number is None or max_member_number < 10:
            new_member_number = 10
        else:
            new_member_number = max_member_number + 1
        
        # Create member profile
        # Convert empty strings to None for optional fields
        cnic_value = data.get('cnic') if data.get('cnic') else None
        email_value = data.get('email') if data.get('email') else None
        
        member = MemberProfile(
            user_id=user.id,
            member_number=new_member_number,
            full_name=data['full_name'],
            phone=data['phone'],
            cnic=cnic_value,  # None if empty
            email=email_value,  # None if empty
            gender=data.get('gender'),
            date_of_birth=dob,
            admission_date=admission_date,
            current_package_id=data.get('package_id') if data.get('package_id') else None,
            trainer_id=data.get('trainer_id') if data.get('trainer_id') else None,
            package_start_date=package_start_date,
            package_expiry_date=package_expiry_date,
            is_frozen=False,
            profile_picture=data.get('profile_picture')
        )
        db.session.add(member)
        db.session.flush()
        
        # Create admission fee transaction (PENDING status)
        # Due date should be the package expiry date if package is assigned, otherwise 7 days from now
        admission_fee = float(data.get('admission_fee', 5000))  # Use provided fee or default to 5000
        discount = float(data.get('discount', 0))
        discount_type = data.get('discount_type', 'fixed')  # 'fixed' or 'percentage'
        trainer_charge = float(data.get('trainer_charge', 0))
        
        # Get package price if package is assigned
        package_price = 0
        if data.get('package_id'):
            package = Package.query.get(data['package_id'])
            if package:
                package_price = float(package.price)
        
        # Calculate total amount for transaction
        # Total = admission_fee + package_price + trainer_charge
        total_amount = admission_fee + package_price + trainer_charge
        
        # Apply discount if any
        if discount > 0:
            if discount_type == 'percentage':
                discount_amount = total_amount * (discount / 100)
            else:
                discount_amount = discount
        else:
            discount_amount = 0
        
        # Final amount after discount
        final_amount = total_amount - discount_amount
        
        due_date = package_expiry_date if package_expiry_date else (datetime.utcnow() + timedelta(days=7))
        
        admission_transaction = Transaction(
            member_id=member.id,
            amount=final_amount,
            transaction_type=TransactionType.ADMISSION,
            status=TransactionStatus.PENDING,
            due_date=due_date,
            trainer_fee=trainer_charge,
            package_price=package_price,
            discount_amount=discount_amount,
            discount_type=discount_type,
            created_at=datetime.utcnow()
        )
        db.session.add(admission_transaction)
        db.session.commit()
        
        member_dict = member.to_dict()
        
        return jsonify(member_dict), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>', methods=['PUT'])
@require_admin
def update_member(member_id):
    """Update a member."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    data = request.get_json()
    
    try:
        # Update member profile fields
        if 'full_name' in data:
            member.full_name = data['full_name']
        if 'phone' in data:
            member.phone = data['phone']
        if 'email' in data:
            member.email = data['email']
        if 'cnic' in data:
            member.cnic = data['cnic']
        if 'address' in data:
            member.address = data['address']
        if 'emergency_contact' in data:
            member.emergency_contact = data['emergency_contact']
        if 'is_frozen' in data:
            member.is_frozen = data['is_frozen']
        if 'profile_picture' in data:
            member.profile_picture = data['profile_picture']
        
        # Handle package assignment (accept both package_id and current_package_id)
        if 'package_id' in data:
            new_package_id = data['package_id'] if data['package_id'] else None
            # If package changed, recalculate dates
            if new_package_id != member.current_package_id and new_package_id:
                package = Package.query.get(new_package_id)
                if package:
                    # Only auto-set dates if not provided
                    if 'package_start_date' not in data or not data['package_start_date']:
                        member.package_start_date = datetime.utcnow()
                    if 'package_expiry_date' not in data or not data['package_expiry_date']:
                        start_date = member.package_start_date or datetime.utcnow()
                        member.package_expiry_date = start_date + timedelta(days=package.duration_days)
            member.current_package_id = new_package_id
        elif 'current_package_id' in data:
            new_package_id = data['current_package_id'] if data['current_package_id'] else None
            # If package changed, recalculate dates
            if new_package_id != member.current_package_id and new_package_id:
                package = Package.query.get(new_package_id)
                if package:
                    # Only auto-set dates if not provided
                    if 'package_start_date' not in data or not data['package_start_date']:
                        member.package_start_date = datetime.utcnow()
                    if 'package_expiry_date' not in data or not data['package_expiry_date']:
                        start_date = member.package_start_date or datetime.utcnow()
                        member.package_expiry_date = start_date + timedelta(days=package.duration_days)
            member.current_package_id = new_package_id
        
        # Handle package dates
        if 'package_start_date' in data:
            try:
                from datetime import datetime as dt
                member.package_start_date = dt.fromisoformat(data['package_start_date'].replace('Z', '+00:00')) if data['package_start_date'] else None
            except:
                pass
        
        if 'package_expiry_date' in data:
            try:
                from datetime import datetime as dt
                member.package_expiry_date = dt.fromisoformat(data['package_expiry_date'].replace('Z', '+00:00')) if data['package_expiry_date'] else None
            except:
                pass
        
        # Handle trainer assignment
        if 'trainer_id' in data:
            member.trainer_id = data['trainer_id'] if data['trainer_id'] else None
        
        if 'gender' in data:
            member.gender = data['gender']
        if 'date_of_birth' in data:
            try:
                from datetime import datetime as dt
                member.date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date() if data['date_of_birth'] else None
            except:
                pass
        
        # Handle admission_date
        if 'admission_date' in data:
            try:
                from datetime import datetime as dt
                member.admission_date = dt.strptime(data['admission_date'], '%Y-%m-%d').date() if data['admission_date'] else None
            except:
                pass
        
        db.session.commit()
        
        member_dict = member.to_dict()
        user = User.query.get(member.user_id)
        if user:
            member_dict['username'] = user.username
        
        return jsonify(member_dict), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>', methods=['DELETE'])
@require_admin
def delete_member(member_id):
    """Delete a member."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    try:
        user = User.query.get(member.user_id)
        db.session.delete(member)
        if user:
            db.session.delete(user)
        db.session.commit()
        
        return jsonify({'message': 'Member deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# TRAINER ROUTES
# ============================================================================

@admin_complete_bp.route('/trainers', methods=['GET'])
@require_admin
def list_trainers():
    """List all trainers with pagination."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    paginated = TrainerProfile.query.paginate(page=page, per_page=per_page, error_out=False)
    
    trainers = []
    for trainer in paginated.items:
        trainer_dict = trainer.to_dict()
        user = User.query.get(trainer.user_id)
        if user:
            trainer_dict['username'] = user.username
        
        # Count assigned members
        assigned_members_count = MemberProfile.query.filter_by(trainer_id=trainer.id).count()
        trainer_dict['assigned_members_count'] = assigned_members_count
        
        trainers.append(trainer_dict)
    
    return jsonify({
        'trainers': trainers,
        'total': paginated.total,
        'page': page,
        'per_page': per_page
    }), 200


@admin_complete_bp.route('/trainers/<trainer_id>', methods=['GET'])
@require_admin
def get_trainer(trainer_id):
    """Get a specific trainer by ID."""
    trainer = TrainerProfile.query.get(trainer_id)
    if not trainer:
        return jsonify({'error': 'Trainer not found'}), 404
    
    trainer_dict = trainer.to_dict()
    user = User.query.get(trainer.user_id)
    if user:
        trainer_dict['username'] = user.username
    
    return jsonify(trainer_dict), 200


@admin_complete_bp.route('/trainers', methods=['POST'])
@require_admin
def create_trainer():
    """Create a new trainer."""
    data = request.get_json()
    
    # Validate required fields (no username/password needed)
    required_fields = ['specialization', 'phone', 'email']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    try:
        # Create user with auto-generated username (email-based)
        auto_username = data['email'].split('@')[0] + '_' + str(uuid.uuid4())[:8]
        auto_password = str(uuid.uuid4())
        
        user = User(
            username=auto_username,
            password_hash=PasswordService.hash_password(auto_password),
            role=UserRole.TRAINER,
            is_active=True
        )
        db.session.add(user)
        db.session.flush()
        
        # Parse date_of_birth if provided
        dob = None
        if data.get('date_of_birth'):
            try:
                from datetime import datetime as dt
                dob = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except:
                dob = None
        
        # Create trainer profile
        trainer = TrainerProfile(
            user_id=user.id,
            specialization=data['specialization'],
            phone=data['phone'],
            email=data['email'],
            full_name=data.get('full_name'),
            gender=data.get('gender'),
            date_of_birth=dob,
            cnic=data.get('cnic'),
            salary_rate=data.get('salary_rate', 0),
            availability=data.get('availability')
        )
        db.session.add(trainer)
        db.session.commit()
        
        trainer_dict = trainer.to_dict()
        
        return jsonify(trainer_dict), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/trainers/<trainer_id>', methods=['PUT'])
@require_admin
def update_trainer(trainer_id):
    """Update a trainer."""
    trainer = TrainerProfile.query.get(trainer_id)
    if not trainer:
        return jsonify({'error': 'Trainer not found'}), 404
    
    data = request.get_json()
    
    try:
        # Update trainer profile fields
        if 'full_name' in data:
            trainer.full_name = data['full_name']
        if 'specialization' in data:
            trainer.specialization = data['specialization']
        if 'phone' in data:
            trainer.phone = data['phone']
        if 'email' in data:
            trainer.email = data['email']
        if 'bio' in data:
            trainer.bio = data['bio']
        if 'gender' in data:
            trainer.gender = data['gender']
        if 'date_of_birth' in data:
            try:
                from datetime import datetime as dt
                trainer.date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date() if data['date_of_birth'] else None
            except:
                pass
        if 'cnic' in data:
            trainer.cnic = data['cnic']
        if 'salary_rate' in data:
            trainer.salary_rate = data['salary_rate']
        if 'availability' in data:
            trainer.availability = data['availability']
        
        db.session.commit()
        
        trainer_dict = trainer.to_dict()
        user = User.query.get(trainer.user_id)
        if user:
            trainer_dict['username'] = user.username
        
        return jsonify(trainer_dict), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/trainers/<trainer_id>/update-fixed', methods=['PUT'])
@require_admin
def update_trainer_fixed(trainer_id):
    """Update a trainer (alias endpoint for frontend compatibility)."""
    return update_trainer(trainer_id)


@admin_complete_bp.route('/trainers/<trainer_id>', methods=['DELETE'])
@require_admin
def delete_trainer(trainer_id):
    """Delete a trainer."""
    trainer = TrainerProfile.query.get(trainer_id)
    if not trainer:
        return jsonify({'error': 'Trainer not found'}), 404
    
    try:
        user = User.query.get(trainer.user_id)
        db.session.delete(trainer)
        if user:
            db.session.delete(user)
        db.session.commit()
        
        return jsonify({'message': 'Trainer deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DASHBOARD METRICS
# ============================================================================

@admin_complete_bp.route('/dashboard/metrics', methods=['GET'])
@require_admin
def get_dashboard_metrics():
    """Get dashboard metrics (without attendance data)."""
    try:
        # Total active members
        total_members = MemberProfile.query.filter_by(is_frozen=False).count()
        
        # Overdue payments - count PENDING transactions where due_date < now
        # EXCLUDE transactions for frozen members
        from datetime import datetime
        now = datetime.utcnow()
        
        # Get all overdue transactions
        overdue_transactions = Transaction.query.filter(
            Transaction.status == TransactionStatus.PENDING,
            Transaction.due_date < now
        ).all()
        
        # Filter out transactions for frozen members
        overdue_payments_count = 0
        for transaction in overdue_transactions:
            member = MemberProfile.query.get(transaction.member_id)
            if member and not member.is_frozen:
                overdue_payments_count += 1
        
        # Since attendance was removed, set these to 0
        daily_attendance_count = 0
        live_floor_count = 0
        
        result = {
            'total_members': total_members,
            'overdue_payments_count': overdue_payments_count,
            'daily_attendance_count': daily_attendance_count,
            'live_floor_count': live_floor_count
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# SETTINGS ROUTES
# ============================================================================

@admin_complete_bp.route('/settings', methods=['GET'])
@require_admin
def get_admin_settings():
    """Get system settings."""
    from models.settings import Settings
    
    try:
        # Get system settings (should be only one record)
        settings = Settings.query.first()
        if not settings:
            # Create default settings if none exist
            settings = Settings(admission_fee=5000)
            db.session.add(settings)
            db.session.commit()
        
        return jsonify({
            'admission_fee': float(settings.admission_fee),
            'package_price': 0,
            'package_duration': 30,
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/settings', methods=['PUT'])
@require_admin
def update_admin_settings():
    """Update system settings."""
    from models.settings import Settings
    
    data = request.get_json()
    
    try:
        # Get or create settings record
        settings = Settings.query.first()
        if not settings:
            settings = Settings()
            db.session.add(settings)
        
        # Update admission_fee if provided
        if 'admission_fee' in data:
            settings.admission_fee = float(data['admission_fee'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Settings updated successfully',
            'admission_fee': float(settings.admission_fee),
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/profile', methods=['PUT'])
@require_admin
def update_admin_profile():
    """Update admin profile."""
    from flask_jwt_extended import get_jwt_identity
    
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    data = request.get_json()
    
    try:
        # Update password if provided
        if 'current_password' in data and 'new_password' in data:
            if not PasswordService.verify_password(data['current_password'], user.password_hash):
                return jsonify({'error': 'Current password is incorrect'}), 400
            
            user.password_hash = PasswordService.hash_password(data['new_password'])
        
        # Update username if provided
        if 'username' in data and data['username'] != user.username:
            existing = User.query.filter_by(username=data['username']).first()
            if existing:
                return jsonify({'error': 'Username already exists'}), 400
            user.username = data['username']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Profile updated successfully',
            'user': user.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



# ============================================================================
# ANALYTICS ENDPOINTS
# ============================================================================

@admin_complete_bp.route('/dashboard/revenue-projection', methods=['GET'])
@require_admin
def get_revenue_projection():
    """Get revenue projection data based on active member packages."""
    try:
        # Get all active members (not frozen) with packages
        active_members = MemberProfile.query.filter(
            MemberProfile.is_frozen == False,
            MemberProfile.current_package_id.isnot(None)
        ).all()
        
        projected_monthly_revenue = 0
        active_packages_count = 0
        now = datetime.now()
        
        for member in active_members:
            # Check if package is still active (not expired or no expiry date set)
            if member.package_expiry_date is None or member.package_expiry_date >= now:
                package = Package.query.get(member.current_package_id)
                if package and package.is_active:
                    # Calculate monthly revenue based on package duration
                    if package.duration_days > 0:
                        monthly_rate = (float(package.price) / package.duration_days) * 30
                        projected_monthly_revenue += monthly_rate
                        active_packages_count += 1
        
        return jsonify({
            'projected_monthly_revenue': round(projected_monthly_revenue, 2),
            'active_packages_count': active_packages_count,
            'current_revenue': round(projected_monthly_revenue, 2),
            'growth_rate': 0
        }), 200
    except Exception as e:
        print(f"Error in get_revenue_projection: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/revenue-trend', methods=['GET'])
@require_admin
def get_revenue_trend():
    """Get revenue trend data."""
    try:
        from datetime import datetime, timedelta
        
        # Get last 6 months of revenue
        months = []
        revenue = []
        
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            months.insert(0, month_date.strftime('%b'))
            
            # Calculate revenue for this month
            month_start = month_date.replace(day=1)
            if i > 0:
                next_month = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            else:
                next_month = datetime.now()
            
            month_revenue = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.status == TransactionStatus.COMPLETED,
                Transaction.paid_date >= month_start,
                Transaction.paid_date < next_month
            ).scalar() or 0
            
            revenue.insert(0, float(month_revenue))
        
        return jsonify({
            'months': months,
            'revenue': revenue
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/daily-revenue', methods=['GET'])
@require_admin
def get_daily_revenue():
    """Get daily revenue data for the last 30 days."""
    try:
        from datetime import datetime, timedelta
        
        # Get last 30 days of revenue
        days = []
        revenue = []
        
        for i in range(30):
            day_date = datetime.now() - timedelta(days=29-i)
            days.append(day_date.strftime('%d %b'))
            
            # Calculate revenue for this day
            day_start = day_date.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            day_revenue = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.status == TransactionStatus.COMPLETED,
                Transaction.paid_date >= day_start,
                Transaction.paid_date < day_end,
                (Transaction.is_reversed == False) | (Transaction.is_reversed.is_(None))  # Exclude reversed transactions
            ).scalar() or 0
            
            revenue.append(float(day_revenue))
        
        return jsonify({
            'days': days,
            'revenue': revenue
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/member-growth', methods=['GET'])
@require_admin
def get_member_growth():
    """Get member growth data."""
    try:
        from datetime import datetime, timedelta
        
        # Get last 6 months of member growth
        months = []
        counts = []
        
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            months.insert(0, month_date.strftime('%b'))
            
            # Count members created up to this month
            month_end = month_date.replace(day=28) + timedelta(days=4)
            month_end = month_end.replace(day=1)
            
            count = User.query.filter(
                User.role == UserRole.MEMBER,
                User.created_at < month_end
            ).count()
            
            counts.insert(0, count)
        
        return jsonify({
            'months': months,
            'counts': counts
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/attendance-trend', methods=['GET'])
@require_admin
def get_attendance_trend():
    """Get attendance trend data (returns empty since attendance was removed)."""
    try:
        from datetime import datetime, timedelta
        
        # Return empty data since attendance was removed
        months = []
        attendance_count = []
        
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            months.insert(0, month_date.strftime('%b'))
            attendance_count.insert(0, 0)
        
        return jsonify({
            'months': months,
            'attendance_count': attendance_count
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/peak-hours', methods=['GET'])
@require_admin
def get_peak_hours():
    """Get peak hours data (returns empty since attendance was removed)."""
    try:
        # Return empty data since attendance was removed
        peak_hours = []
        
        for hour in range(6, 23):  # 6 AM to 10 PM
            peak_hours.append({
                'hour': hour,
                'count': 0
            })
        
        return jsonify({
            'peak_hours': peak_hours
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# FINANCE ENDPOINTS (for AdminFinance page)
# ============================================================================

@admin_complete_bp.route('/finance/member-payments-fixed', methods=['GET'])
@require_admin
def get_member_payments_fixed():
    """Get all transactions with proper status calculation."""
    try:
        # Get all transactions
        transactions = Transaction.query.all()
        
        payments = []
        for transaction in transactions:
            try:
                member = MemberProfile.query.get(transaction.member_id)
                if not member:
                    continue
                
                user = User.query.get(member.user_id)
                username = user.username if user else 'Unknown'
                
                # Calculate status based on transaction state and dates
                status = transaction.status.value
                
                # If not completed, check if overdue
                if status != 'COMPLETED' and transaction.due_date:
                    if datetime.utcnow() > transaction.due_date:
                        status = 'OVERDUE'
                
                # Safely get trainer_fee and discount_amount
                trainer_fee = 0
                discount_amount = 0
                package_price = 0
                try:
                    trainer_fee = float(transaction.trainer_fee) if transaction.trainer_fee else 0
                except (TypeError, ValueError):
                    trainer_fee = 0
                
                try:
                    discount_amount = float(transaction.discount_amount) if transaction.discount_amount else 0
                except (TypeError, ValueError):
                    discount_amount = 0
                
                try:
                    package_price = float(transaction.package_price) if transaction.package_price else 0
                except (TypeError, ValueError):
                    package_price = 0
                
                payments.append({
                    'id': transaction.id,
                    'member_id': transaction.member_id,
                    'username': username,
                    'full_name': member.full_name,
                    'phone': member.phone,
                    'amount': float(transaction.amount),
                    'transaction_type': transaction.transaction_type.value if hasattr(transaction.transaction_type, 'value') else str(transaction.transaction_type),
                    'status': status,
                    'due_date': transaction.due_date.isoformat() + 'Z' if transaction.due_date else None,
                    'paid_date': transaction.paid_date.isoformat() + 'Z' if transaction.paid_date else None,
                    'created_at': transaction.created_at.isoformat() + 'Z',
                    'total_monthly_payment': float(transaction.amount),
                    'trainer_fee': trainer_fee,
                    'package_price': package_price,
                    'discount_amount': discount_amount,
                    'discount_type': transaction.discount_type or 'fixed',
                    'is_reversed': getattr(transaction, 'is_reversed', False),
                    'reversed_at': transaction.reversed_at.isoformat() + 'Z' if hasattr(transaction, 'reversed_at') and transaction.reversed_at else None,
                    'reversed_by': getattr(transaction, 'reversed_by', None)
                })
            except Exception as item_error:
                continue
        
        return jsonify({'payments': payments}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/transactions/<transaction_id>/mark-paid', methods=['POST'])
@require_admin
def mark_transaction_paid(transaction_id):
    """Mark a transaction as paid and create next month's transaction."""
    try:
        transaction = Transaction.query.get(transaction_id)
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404
        
        # Get member info
        member = MemberProfile.query.get(transaction.member_id)
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        # Update transaction status and paid date
        transaction.status = TransactionStatus.COMPLETED
        transaction.paid_date = datetime.utcnow()
        
        # If this is an admission fee, mark admission_fee_paid on member
        if transaction.transaction_type.value == 'ADMISSION':
            if member:
                member.admission_fee_paid = True
        
        # Create next month's transaction automatically (only for MONTHLY type or if member has active package)
        # Don't create for frozen members
        if not member.is_frozen and member.current_package_id:
            # Get package details
            package = Package.query.get(member.current_package_id)
            if package and package.is_active:
                # Calculate next month's due date (30 days from current due date)
                next_due_date = transaction.due_date + timedelta(days=30) if transaction.due_date else datetime.utcnow() + timedelta(days=30)
                
                # Get trainer fee if trainer is assigned
                trainer_fee = 0
                if member.trainer_id:
                    trainer = TrainerProfile.query.get(member.trainer_id)
                    if trainer:
                        trainer_fee = float(trainer.salary_rate) if trainer.salary_rate else 0
                
                # Calculate total amount (package price + trainer fee)
                package_price = float(package.price) if package.price else 0
                total_amount = package_price + trainer_fee
                
                # Create new transaction for next month
                new_transaction = Transaction(
                    member_id=member.id,
                    amount=total_amount,
                    transaction_type=TransactionType.PAYMENT,
                    status=TransactionStatus.PENDING,
                    due_date=next_due_date,
                    trainer_fee=trainer_fee,
                    package_price=package_price,
                    discount_amount=0,
                    discount_type='fixed',
                    created_at=datetime.utcnow()
                )
                db.session.add(new_transaction)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Transaction marked as paid and next month transaction created',
            'transaction': transaction.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/transactions/<transaction_id>/reverse-payment', methods=['POST'])
@require_admin
def reverse_transaction_payment(transaction_id):
    """Reverse a completed payment within 24 hours."""
    try:
        from flask import g
        
        transaction = Transaction.query.get(transaction_id)
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404
        
        # Check if transaction can be reversed
        if not transaction.can_reverse():
            if transaction.is_reversed:
                return jsonify({'error': 'Transaction has already been reversed'}), 400
            elif transaction.status != TransactionStatus.COMPLETED:
                return jsonify({'error': 'Only completed transactions can be reversed'}), 400
            elif not transaction.paid_date:
                return jsonify({'error': 'Transaction has no payment date'}), 400
            else:
                return jsonify({'error': 'Transaction can only be reversed within 24 hours of payment'}), 400
        
        # Get member info
        member = MemberProfile.query.get(transaction.member_id)
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        # Mark transaction as reversed
        transaction.status = TransactionStatus.PENDING
        transaction.is_reversed = True
        transaction.reversed_at = datetime.utcnow()
        transaction.reversed_by = g.current_user.id if hasattr(g, 'current_user') else None
        
        # Clear paid_date to indicate it's no longer paid
        original_paid_date = transaction.paid_date
        transaction.paid_date = None
        
        # If this was an admission fee, reverse the admission_fee_paid flag
        if transaction.transaction_type.value == 'ADMISSION':
            member.admission_fee_paid = False
        
        # Find and delete the next month's transaction that was auto-created
        # Look for transactions created after this payment with due_date 30 days later
        if original_paid_date and not member.is_frozen:
            next_due_date = transaction.due_date + timedelta(days=30) if transaction.due_date else None
            if next_due_date:
                # Find auto-created next transaction
                next_transaction = Transaction.query.filter_by(
                    member_id=member.id,
                    status=TransactionStatus.PENDING,
                    due_date=next_due_date
                ).filter(
                    Transaction.created_at > original_paid_date
                ).first()
                
                if next_transaction:
                    db.session.delete(next_transaction)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Payment reversed successfully. Transaction status reset to PENDING.',
            'transaction': transaction.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# TRAINERS ENDPOINTS (for AdminTrainers page)
# ============================================================================

@admin_complete_bp.route('/trainers-fixed', methods=['GET'])
@require_admin
def list_trainers_fixed():
    """List all trainers (alias for /trainers endpoint)."""
    return list_trainers()


# ============================================================================
# ADMIN SETTINGS ENDPOINTS
# ============================================================================

@admin_complete_bp.route('/change-password', methods=['POST'])
@jwt_required()
@require_admin
def change_password():
    """Change admin password."""
    try:
        from flask_jwt_extended import get_jwt_identity
        
        data = request.get_json()
        
        if not data or not data.get('current_password') or not data.get('new_password'):
            return jsonify({'error': 'Missing required fields'}), 400
        
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        # Validate new password length
        if len(new_password) < 6:
            return jsonify({'error': 'Password must be at least 6 characters'}), 400
        
        # Get current user
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Verify current password
        if not PasswordService.verify_password(current_password, user.password_hash):
            return jsonify({'error': 'Current password is incorrect'}), 401
        
        # Update password
        user.password_hash = PasswordService.hash_password(new_password)
        db.session.commit()
        
        return jsonify({'message': 'Password changed successfully'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error in change_password: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
